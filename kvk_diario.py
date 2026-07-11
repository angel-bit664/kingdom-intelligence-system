import pandas as pd
import discord
from discord.ext import commands
from discord import app_commands
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import zipfile

# ===== CONFIG =====
PODER_MINIMO = 0 # 0 = Todos los jugadores | 30_000_000 = Solo >30M
MERITOS_ESPERADOS_POR_DIA = 100_000 # Meta: 100k méritos diarios
# ==================

COLORES = {
    'azul_oscuro': '1F4E78',
    'azul_claro': '4472C4',
    'verde': '70AD47',
    'rojo': 'C00000',
    'amarillo': 'FFC000',
    'naranja': 'FF6B35',
    'gris': 'D9D9D9',
    'blanco': 'FFFFFF'
}

# --- FUNCIONES AUXILIARES ---

def limpiar_numero(serie):
    """Convierte strings como '1.5M', '200K' a números flotantes"""
    return (serie.astype(str)
              .str.replace(',', '', regex=False)
              .str.replace(' ', '', regex=False)
              .str.replace('M', 'e6', case=False, regex=False)
              .str.replace('K', 'e3', case=False, regex=False)
              .str.replace('B', 'e9', case=False, regex=False)
              .pipe(pd.to_numeric, errors='coerce').fillna(0))

def detectar_columna(df, posibles):
    """Detecta columna aunque tenga espacios, mayúsculas o palabras extra (ej: nombre_de_personaje)"""
    df_cols = {str(col).lower().strip(): col for col in df.columns}
    for posible in posibles:
        if posible.lower() in df_cols:
            return df_cols[posible.lower()]
    for posible in posibles:
        for col_lower, col_real in df_cols.items():
            if posible.lower() in col_lower:
                return col_real
    raise ValueError(f"❌ No encontré columna. Buscaba: {posibles}. Tengo: {list(df.columns)}")

def ajustar_ancho_columnas(ws, min_ancho=10, max_ancho=40):
    """Ajusta el ancho de las columnas automáticamente según el contenido"""
    for col in ws.columns:
        max_length = 0
        columna = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ajustado = max_length + 2
        ws.column_dimensions[columna].width = min(max(ajustado, min_ancho), max_ancho)

# --- LÓGICA PRINCIPAL ---

async def procesar_kvk(archivos_discord):
    # 1. LEER DESDE MEMORIA RAM (Ideal para Railway)
    dfs = []
    nombres_archivos = []
    
    for archivo in archivos_discord:
        file_bytes = await archivo.read()
        if archivo.filename.endswith('.zip'):
            with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as zip_ref:
                for name in zip_ref.namelist():
                    if name.endswith('.xlsx'):
                        excel_bytes = zip_ref.read(name)
                        dfs.append(pd.read_excel(io.BytesIO(excel_bytes)))
                        nombres_archivos.append(name.replace('.xlsx', ''))
        elif archivo.filename.endswith('.xlsx'):
            dfs.append(pd.read_excel(io.BytesIO(file_bytes)))
            nombres_archivos.append(archivo.filename.replace('.xlsx', ''))

    if len(dfs) < 2:
        raise ValueError("❌ Necesitas mínimo 2 días de KVK en Excel o 1 ZIP con varios")

    # 2. LIMPIAR COLUMNAS
    for i in range(len(dfs)):
        dfs[i].columns = dfs[i].columns.str.lower().str.strip().str.replace(' ', '_')

    df_inicial = dfs[0].copy()
    df_final = dfs[-1].copy()
    dia_actual = len(dfs)

    # 3. DETECTAR COLUMNAS
    col_nombre = detectar_columna(df_final, ['nombre', 'name', 'jugador', 'player', 'nombre_de_personaje'])
    col_poder = detectar_columna(df_final, ['poder', 'power', 'poder_actual'])
    col_meritos = detectar_columna(df_final, ['meritos', 'méritos', 'merits', 'honor'])

    # 4. LIMPIAR NÚMEROS
    for df in [df_inicial, df_final]:
        df[col_poder] = limpiar_numero(df[col_poder])
        df[col_meritos] = limpiar_numero(df[col_meritos])

    # 5. FILTRAR
    if PODER_MINIMO > 0:
        df_inicial = df_inicial[df_inicial[col_poder] >= PODER_MINIMO].copy()
        df_final = df_final[df_final[col_poder] >= PODER_MINIMO].copy()
    if len(df_final) == 0:
        raise ValueError(f"❌ No hay jugadores con poder > {PODER_MINIMO:,}")

    # 6. DETECTAR ALTAS/BAJAS Y MERGE
    jugadores_inicial = set(df_inicial[col_nombre].dropna().astype(str))
    jugadores_final = set(df_final[col_nombre].dropna().astype(str))
    nuevos = jugadores_final - jugadores_inicial
    bajas = jugadores_inicial - jugadores_final

    df = df_final.merge(df_inicial, on=col_nombre, how='outer', suffixes=('', '_old'))
    df = df.rename(columns={
        f'{col_poder}': 'poder_actual', f'{col_poder}_old': 'poder_inicial',
        f'{col_meritos}': 'meritos_final', f'{col_meritos}_old': 'meritos_inicial'
    })
    
    for col in ['poder_actual', 'poder_inicial', 'meritos_final', 'meritos_inicial']:
        df[col] = df[col].fillna(0)
    if PODER_MINIMO > 0:
        df = df[df['poder_actual'] >= PODER_MINIMO].copy()

    # 7. CÁLCULOS
    df['cambio_poder'] = df['poder_actual'] - df['poder_inicial']
    df['cambio_poder_pct'] = ((df['poder_actual'] / df['poder_inicial'].replace(0, 1)) - 1) * 100
    df['meta_acumulada'] = df['poder_inicial'] * (1 + 0.017 * dia_actual)
    df['porcentaje_avance'] = ((df['poder_actual'] / df['meta_acumulada'].replace(0, 1)) - 1) * 100
    df['meritos_ganados'] = df['meritos_final'] - df['meritos_inicial']
    df['meritos_por_dia'] = df['meritos_ganados'] / dia_actual
    df['poder_maximo'] = df[['poder_actual', 'poder_inicial']].max(axis=1)
    df['pct_meritos_vs_poder'] = (df['meritos_ganados'] / df['poder_maximo'].replace(0, 1)) * 100
    df['meritos_por_semana'] = df['meritos_ganados'] / (dia_actual / 7)
    df['eficiencia_meritos'] = df['meritos_ganados'] / (df['poder_maximo'] / 1_000_000).replace(0, 1)
    df['meta_meritos'] = MERITOS_ESPERADOS_POR_DIA * dia_actual
    df['pct_meta_meritos'] = ((df['meritos_ganados'] / df['meta_meritos'].replace(0, 1)) - 1) * 100

    # 8. ESTADOS
    df['estado'] = '🟡 Normal'
    df.loc[df[col_nombre].isin(nuevos), 'estado'] = '🆕 NUEVO'
    df.loc[df[col_nombre].isin(bajas), 'estado'] = '❌ BAJA'
    df.loc[(df['estado']!= '❌ BAJA') & (df['porcentaje_avance'] >= 0), 'estado'] = '🟢 Cumple Meta'
    df.loc[(df['estado']!= '❌ BAJA') & (df['poder_actual'] >= 150_000_000) & (df['cambio_poder'] < 0), 'estado'] = '🔴 Ballena Muerta'
    df.loc[(df['estado']!= '❌ BAJA') & (df['meritos_ganados'] < 500_000) & (df['poder_actual'] > 0), 'estado'] = '👻 Fantasma'
    df.loc[(df['estado']!= '❌ BAJA') & (df['poder_actual'] >= 50_000_000) & (df['porcentaje_avance'] < -5), 'estado'] = '⚠️ Riesgo Kick'
    df.loc[(df['estado']!= '❌ BAJA') & (df['pct_meta_meritos'] < -50), 'estado'] = '📉 Sin Méritos'

    # 9. RANKINGS
    df_final_rank = df_final[[col_nombre, col_poder]].copy()
    df_final_rank['rank_actual'] = df_final_rank[col_poder].rank(ascending=False, method='min')
    df_inicial_rank = df_inicial[[col_nombre, col_poder]].copy()
    df_inicial_rank['rank_inicial'] = df_inicial_rank[col_poder].rank(ascending=False, method='min')

    df = df.merge(df_final_rank[[col_nombre, 'rank_actual']], on=col_nombre, how='left')
    df = df.merge(df_inicial_rank[[col_nombre, 'rank_inicial']], on=col_nombre, how='left')
    df['cambio_rank'] = df['rank_inicial'].fillna(df['rank_actual']) - df['rank_actual']

    # 10. MÉTRICAS GLOBALES
    df_activos = df[df['estado']!= '❌ BAJA'].copy()
    total = len(df_activos)
    cumplen = len(df_activos[df_activos['porcentaje_avance'] >= 0])
    pct_cumple = (cumplen / total) * 100 if total > 0 else 0
    poder_ganado = df_activos[df_activos['cambio_poder'] > 0]['cambio_poder'].sum()
    poder_perdido = abs(df_activos[df_activos['cambio_poder'] < 0]['cambio_poder'].sum())
    fantasmas = len(df_activos[df_activos['estado'] == '👻 Fantasma'])
    ballenas_muertas = len(df_activos[df_activos['estado'] == '🔴 Ballena Muerta'])
    riesgo_kick = len(df_activos[df_activos['estado'] == '⚠️ Riesgo Kick'])
    sin_meritos = len(df_activos[df_activos['estado'] == '📉 Sin Méritos'])
    meritos_totales = df_activos['meritos_ganados'].sum()

    # --- CREAR EXCEL EN MEMORIA ---
    wb = Workbook()
    border_thick = Border(left=Side(style='medium', color=COLORES['azul_oscuro']), right=Side(style='medium', color=COLORES['azul_oscuro']), top=Side(style='medium', color=COLORES['azul_oscuro']), bottom=Side(style='medium', color=COLORES['azul_oscuro']))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    titulo_style = Font(bold=True, size=22, color=COLORES['blanco'], name='Calibri')
    header_style = Font(bold=True, size=11, color=COLORES['blanco'], name='Calibri')

    # ===== HOJA 1: DASHBOARD =====
    ws_dash = wb.active
    ws_dash.title = "EXECUTIVE DASHBOARD"
    
    ws_dash.merge_cells('A1:P3')
    ws_dash['A1'] = f"⚔️ REPORTE EJECUTIVO KVK | DÍA {dia_actual}"
    ws_dash['A1'].font = titulo_style
    ws_dash['A1'].fill = PatternFill("solid", fgColor=COLORES['azul_oscuro'])
    ws_dash['A1'].alignment = center
    ws_dash['A1'].border = border_thick

    ws_dash.merge_cells('A4:P4')
    filtro_txt = f"Filtro: >{PODER_MINIMO/1e6:.0f}M" if PODER_MINIMO > 0 else "Todos los jugadores"
    ws_dash['A4'] = f"Período: {nombres_archivos[0]} → {nombres_archivos[-1]} | {filtro_txt}"
    ws_dash['A4'].font = Font(size=11, italic=True, color=COLORES['azul_oscuro'])
    ws_dash['A4'].fill = PatternFill("solid", fgColor=COLORES['gris'])
    ws_dash['A4'].alignment = center

    # KPIs
    kpi_data = [
        ("MIEMBROS ACTIVOS", f"{total}", f"+{len(nuevos)} -{len(bajas)}", COLORES['azul_claro']),
        ("TASA CUMPLIMIENTO", f"{pct_cumple:.1f}%", f"{cumplen}/{total}", COLORES['verde'] if pct_cumple >= 50 else COLORES['rojo']),
        ("PODER GANADO", f"{poder_ganado/1e9:.2f}B", "Total Neto Positivo", COLORES['verde']),
        ("PODER PERDIDO", f"{poder_perdido/1e9:.2f}B", "Total Neto Negativo", COLORES['rojo']),
        ("MÉRITOS/DÍA", f"{meritos_totales/dia_actual/1e6:.1f}M", f"Meta: {MERITOS_ESPERADOS_POR_DIA/1e3:.0f}K", COLORES['naranja']),
        ("FANTASMAS", f"{fantasmas}", f"{fantasmas/total*100:.0f}% del total", COLORES['amarillo'])
    ]

    for i, (titulo, valor, sub, color_fill) in enumerate(kpi_data):
        col = 1 + (i * 3)
        ws_dash.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+2)
        ws_dash.cell(6, col, titulo).font = Font(bold=True, size=10, color=COLORES['azul_oscuro'])
        ws_dash.cell(6, col).alignment = center
        
        ws_dash.merge_cells(start_row=7, start_column=col, end_row=8, end_column=col+2)
        ws_dash.cell(7, col, valor).font = Font(bold=True, size=24, color=COLORES['blanco'])
        ws_dash.cell(7, col).fill = PatternFill("solid", fgColor=color_fill)
        ws_dash.cell(7, col).alignment = center
        ws_dash.cell(7, col).border = border_thick
        
        ws_dash.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+2)
        ws_dash.cell(9, col, sub).font = Font(size=9, color=COLORES['azul_oscuro'])
        ws_dash.cell(9, col).alignment = center

    # TOP 5 MOVERS (Con gráfica de barras)
    ws_dash['A11'] = "📈 TOP 5 CRECIMIENTO (Poder)"
    ws_dash['A11'].font = header_style
    ws_dash['A11'].fill = PatternFill("solid", fgColor=COLORES['verde'])
    ws_dash.merge_cells('A11:D11')

    headers_top = ['Jugador', 'Poder Ganado', '% Avance', 'Méritos']
    for i, h in enumerate(headers_top):
        cell = ws_dash.cell(12, i+1, h)
        cell.font = Font(bold=True, color=COLORES['blanco'])
        cell.fill = PatternFill("solid", fgColor=COLORES['azul_claro'])
        cell.alignment = center

    top_crecimiento = df_activos.nlargest(5, 'cambio_poder')
    for i, (_, row) in enumerate(top_crecimiento.iterrows()):
        ws_dash.cell(13 + i, 1, str(row[col_nombre])[:25])
        ws_dash.cell(13 + i, 2, row['cambio_poder'])
        ws_dash.cell(13 + i, 2).number_format = '#,##0'
        ws_dash.cell(13 + i, 3, row['porcentaje_avance'] / 100)
        ws_dash.cell(13 + i, 3).number_format = '0.0%'
        ws_dash.cell(13 + i, 4, row['meritos_ganados'])
        ws_dash.cell(13 + i, 4).number_format = '#,##0'

    # Crear Gráfica de Barras
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top 5 Crecimiento (Millones)"
    bar.y_axis.title = "Jugador"
    bar.x_axis.title = "Poder Ganado"
    data_ref = Reference(ws_dash, min_col=2, max_col=2, min_row=12, max_row=17)
    cats_ref = Reference(ws_dash, min_col=1, max_col=1, min_row=13, max_row=17)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.shape = 4
    bar.width = 18
    bar.height = 10
    ws_dash.add_chart(bar, "F11")

    ajustar_ancho_columnas(ws_dash, max_ancho=25)

    # ===== HOJA 2: RANKING GENERAL =====
    ws_ranking = wb.create_sheet("RANKING GENERAL")
    ws_ranking['A1'] = f"🏆 RANKING GENERAL - TODOS LOS JUGADORES"
    ws_ranking['A1'].font = titulo_style
    ws_ranking['A1'].fill = PatternFill("solid", fgColor=COLORES['azul_oscuro'])
    ws_ranking['A1'].alignment = center
    ws_ranking.merge_cells('A1:I1')

    df_ranking = df_activos.sort_values('pct_meritos_vs_poder', ascending=False).copy()
    df_ranking['posicion'] = range(1, len(df_ranking) + 1)

    headers_ranking = ['Pos', 'Jugador', 'Poder Máx', 'Méritos Ganados', '% Méritos/Poder', 'Méritos/Semana', 'Efic. Méritos', 'Estado', 'Cambio Rank']
    for i, h in enumerate(headers_ranking):
        cell = ws_ranking.cell(3, i+1, h)
        cell.font = header_style
        cell.fill = PatternFill("solid", fgColor=COLORES['azul_claro'])
        cell.alignment = center
        cell.border = border_thick

    row_start = 4
    for idx, (_, row) in enumerate(df_ranking.iterrows()):
        ws_ranking.cell(row_start + idx, 1, int(row['posicion']))
        ws_ranking.cell(row_start + idx, 2, str(row[col_nombre])[:30])
        ws_ranking.cell(row_start + idx, 3, row['poder_maximo']).number_format = '#,##0'
        ws_ranking.cell(row_start + idx, 4, row['meritos_ganados']).number_format = '#,##0'
        ws_ranking.cell(row_start + idx, 5, row['pct_meritos_vs_poder'] / 100).number_format = '0.00%'
        ws_ranking.cell(row_start + idx, 6, row['meritos_por_semana']).number_format = '#,##0'
        ws_ranking.cell(row_start + idx, 7, row['eficiencia_meritos']).number_format = '0.00'
        ws_ranking.cell(row_start + idx, 8, str(row['estado']))
        ws_ranking.cell(row_start + idx, 9, int(row['cambio_rank']) if pd.notna(row['cambio_rank']) else 0)
        
        for col in range(1, 10):
            ws_ranking.cell(row_start + idx, col).alignment = Alignment(horizontal="center")
            if idx < 3:
                ws_ranking.cell(row_start + idx, col).fill = PatternFill("solid", fgColor='FFD700')

    ws_ranking.auto_filter.ref = f"A3:I{row_start + len(df_ranking) - 1}"
    ws_ranking.freeze_panes = 'A4'
    ajustar_ancho_columnas(ws_ranking)

    # ===== HOJAS DE PROBLEMAS (FANTASMAS, ETC) =====
    hojas_categorias = [
        ("FANTASMAS", '👻 Fantasma'), ("BALLENAS MUERTAS", '🔴 Ballena Muerta'), 
        ("RIESGO KICK", '⚠️ Riesgo Kick'), ("SIN MÉRITOS", '📉 Sin Méritos')
    ]

    for nombre_hoja, estado_buscar in hojas_categorias:
        df_cat = df_activos[df_activos['estado'] == estado_buscar].sort_values('cambio_poder', ascending=True)
        ws = wb.create_sheet(nombre_hoja)
        
        ws['A1'] = f"⚠️ {nombre_hoja} - {len(df_cat)} JUGADORES"
        ws['A1'].font = titulo_style
        ws['A1'].fill = PatternFill("solid", fgColor=COLORES['rojo'])
        ws['A1'].alignment = center
        ws.merge_cells('A1:F1')

        if len(df_cat) > 0:
            headers_cat = ['Nombre', 'Poder Actual', 'Cambio Poder', '% vs Meta', 'Méritos', 'Eficiencia']
            for i, h in enumerate(headers_cat):
                cell = ws.cell(3, i+1, h)
                cell.font = header_style
                cell.fill = PatternFill("solid", fgColor=COLORES['rojo'])
                cell.alignment = center

            for idx, (_, row) in enumerate(df_cat.iterrows()):
                ws.cell(4 + idx, 1, str(row[col_nombre])[:30])
                ws.cell(4 + idx, 2, row['poder_actual']).number_format = '#,##0'
                ws.cell(4 + idx, 3, row['cambio_poder']).number_format = '#,##0'
                ws.cell(4 + idx, 4, row['porcentaje_avance'] / 100).number_format = '0.0%'
                ws.cell(4 + idx, 5, row['meritos_ganados']).number_format = '#,##0'
                ws.cell(4 + idx, 6, row['eficiencia_meritos']).number_format = '0.00'
                for c in range(1, 7):
                    ws.cell(4 + idx, c).alignment = Alignment(horizontal="center")
        ajustar_ancho_columnas(ws)

    # GUARDAR EN MEMORIA RAM
    excel_ram = io.BytesIO()
    wb.save(excel_ram)
    excel_ram.seek(0)

    # RETORNAR ARCHIVO Y DATOS PARA EL MENSAJE
    top_3_crecimiento = df_activos.nlargest(3, 'cambio_poder')
    datos_mensaje = {
        'dia_actual': dia_actual, 'total': total, 'nuevos': len(nuevos), 'bajas': len(bajas),
        'cumplen': cumplen, 'pct_cumple': pct_cumple, 'poder_ganado': poder_ganado,
        'poder_perdido': poder_perdido, 'fantasmas': fantasmas, 'ballenas_muertas': ballenas_muertas,
        'riesgo_kick': riesgo_kick, 'sin_meritos': sin_meritos, 'meritos_totales': meritos_totales,
        'top_3': [(str(row[col_nombre])[:20], row['cambio_poder']) for _, row in top_3_crecimiento.iterrows()],
        'col_nombre': col_nombre
    }
    return excel_ram, datos_mensaje

# --- COMANDO DISCORD ---

class KVKBot(commands.Bot):
    def __init__(self):
        super().__init__(command_prefix="!", intents=discord.Intents.all())

bot = KVKBot()

@bot.event
async def on_ready():
    print(f'✅ Bot conectado como {bot.user}')
    try:
        synced = await bot.tree.sync()
        print(f'✅ Comandos slash sincronizados: {len(synced)}')
    except Exception as e:
        print(f'Error sincronizando: {e}')

@bot.tree.command(name="kvk", description="Genera reporte ejecutivo de KVK subiendo Excels o ZIPs")
@app_commands.describe(archivos="Sube aquí tus archivos de KVK (Excel o ZIP)")
async def kvk_slash(interaction: discord.Interaction, archivos: list[discord.Attachment]):
    # Deferrimos porque Railway puede tardar unos segundos en procesar
    await interaction.response.defer(thinking=True)
    
    try:
        # Validar que sean Excels o Zips
        for arch in archivos:
            if not (arch.filename.endswith('.xlsx') or arch.filename.endswith('.zip')):
                await interaction.followup.send("❌ Todos los archivos deben ser `.xlsx` o `.zip`.")
                return

        # Procesar
        excel_archivo, datos = await procesar_kvk(archivos)
        
        # Calcular Neto CORRECTAMENTE
        neto = datos['poder_ganado'] - datos['poder_perdido']
        filtro_txt = f">{PODER_MINIMO/1e6:.0f}M" if PODER_MINIMO > 0 else ">0M"

        # Crear Embed
        if datos['pct_cumple'] >= 70: color = discord.Color.green()
        elif datos['pct_cumple'] >= 40: color = discord.Color.yellow()
        else: color = discord.Color.red()

        embed = discord.Embed(
            title=f"⚔️ REPORTE EJECUTIVO KVK | DÍA {datos['dia_actual']}",
            description=f"**{'🟢 ESTABLE' if datos['pct_cumple'] >= 70 else '🟡 ALERTA' if datos['pct_cumple'] >= 40 else '🔴 CRÍTICO'}** | Cumplimiento: {datos['pct_cumple']:.1f}% | Filtro: {filtro_txt}",
            color=color
        )

        embed.add_field(name="📊 RESUMEN EJECUTIVO", value=f"👥 {datos['total']} Miembros Activos\n🆕 {datos['nuevos']} Nuevos\n❌ {datos['bajas']} Bajas\n🟢 {datos['cumplen']} Cumplen Meta ({datos['pct_cumple']:.0f}%)", inline=True)
        embed.add_field(name="💰 BALANCE DE PODER", value=f"📈 +{datos['poder_ganado']/1e9:.2f}B Ganado\n📉 -{datos['poder_perdido']/1e9:.2f}B Perdido\n⚖️ {neto/1e9:.2f}B Neto", inline=True)
        embed.add_field(name="⚠️ ALERTAS CRÍTICAS", value=f"👻 {datos['fantasmas']} Fantasmas\n🐋 {datos['ballenas_muertas']} Ballenas Muertas\n⚠️ {datos['riesgo_kick']} Riesgo Kick\n📉 {datos['sin_meritos']} Sin Méritos", inline=True)
        
        top_text = ""
        for i, (nombre, poder) in enumerate(datos['top_3']):
            top_text += f"**{i+1}.** {nombre} (+{poder/1e6:.1f}M)\n"
        embed.add_field(name="🏆 TOP 3 CRECIMIENTO", value=top_text, inline=False)
        
        embed.add_field(name="📁 ARCHIVO EXCEL PROFESIONAL", value="✅ Executive Dashboard\n✅ Movimientos (Altas/Bajas)\n✅ Detalle Completo con Condicionales\n✅ Fantasmas / Ballenas / Riesgo Kick", inline=True)
        embed.set_footer(text=f"Día {datos['dia_actual']}/7 | Méritos Totales: {datos['meritos_totales']/1e6:.1f}M | Filtro: {filtro_txt}")

        # Enviar
        archivo_discord = discord.File(fp=excel_archivo, filename=f"Reporte_KVK_Dia_{datos['dia_actual']}.xlsx")
        await interaction.followup.send(embed=embed, file=archivo_discord)

    except ValueError as e:
        await interaction.followup.send(str(e), ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"⚠️ Ocurrió un error inesperado en el servidor: `{str(e)}`", ephemeral=True)

# --- EJECUTAR BOT ---
import os
from dotenv import load_dotenv

load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')
if not TOKEN:
    print("❌ No se encontró DISCORD_TOKEN en las variables de entorno")
else:
    bot.run(TOKEN)
